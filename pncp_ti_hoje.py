import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
DATA_FINAL = datetime.now().strftime("%Y%m%d")  # hoje AAAAMMDD
MAX_TITENS_TI = 100                            # quantos itens TI você quer
TAMANHO_PAGINA = 10                            # 10 funciona (você testou); 100 dá erro no PNCP
MAX_RAW = 200                                  # amostra sem filtro para inspeção

URL_PROPOSTA = "https://pncp.gov.br/api/consulta/v1/contratacoes/proposta"

# Palavras-chave TI/software (mais fortes)
KEYWORDS_TI = [
    "software", "sistema", "aplicativo", "app", "web", "mobile",
    "desenvolvimento", "manutenção de software", "manutencao de software",
    "fábrica de software", "fabrica de software",
    "ti", "t.i.", "tecnologia da informação", "tecnologia da informacao",
    "infraestrutura de ti", "suporte ti", "service desk", "help desk",
    "rede", "servidor", "cloud", "nuvem",
    "banco de dados", "sql", "postgres", "oracle",
    "api", "integração", "integracao",
    "segurança da informação", "seguranca da informacao", "lgpd",
    "sistema informatizado", "sistemas de informação", "sistemas de informacao",
    "erp", "bi", "dashboard", "etl", "data warehouse",
    "gestão de acessos", "gestao de acessos",
    "cibersegurança", "ciberseguranca",
    "monitoramento", "observabilidade"
]

def pncp_search_link(numero_controle: str) -> str:
    """Link pesquisável no PNCP usando numeroControlePNCP."""
    if not numero_controle:
        return ""
    return f"https://pncp.gov.br/app/editais?q={numero_controle}"

def is_ti(texto: str) -> bool:
    if not texto:
        return False
    t = texto.lower()
    return any(k in t for k in KEYWORDS_TI)

def extract_list(payload):
    """Extrai a lista de itens do JSON (varia por endpoint/versão)."""
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        if isinstance(payload.get("data"), list):
            return payload["data"]
        for v in payload.values():
            if isinstance(v, list):
                return v
    return []

def get_objeto(it: dict) -> str:
    """Tenta pegar o objeto por vários nomes possíveis."""
    return (
        it.get("objeto")
        or it.get("descricaoObjeto")
        or it.get("objetoCompra")
        or it.get("objetoContratacao")
        or it.get("descricao")
        or ""
    )

def get_orgao(it: dict) -> str:
    org = it.get("orgaoEntidade") or {}
    return (
        org.get("razaoSocial")
        or it.get("nomeOrgaoEntidade")
        or it.get("orgao")
        or ""
    )

def get_valor(it: dict):
    return it.get("valorTotalEstimado") or it.get("valorEstimado") or it.get("valorGlobal") or ""

def get_numero_controle(it: dict) -> str:
    return it.get("numeroControlePNCP") or it.get("numeroControlePncp") or it.get("idContratacaoPncp") or ""

def request_json(url, params):
    r = requests.get(url, params=params, headers={"Accept": "application/json"}, timeout=60)

    if r.status_code == 204:
        return "END", None, r.url, ""

    if r.status_code != 200:
        msg = ""
        try:
            msg = r.json().get("message", "")
        except:
            msg = r.text[:300]
        return f"ERR_{r.status_code}", None, r.url, msg

    return "OK", r.json(), r.url, ""

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 90)

def write_sheet(ws, headers, rows, link_col_name=None):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # deixa link clicável
    if link_col_name and link_col_name in headers:
        link_col = headers.index(link_col_name) + 1
        for r in range(2, ws.max_row + 1):
            link = ws.cell(row=r, column=link_col).value
            if link:
                ws.cell(row=r, column=link_col).hyperlink = link
                ws.cell(row=r, column=link_col).style = "Hyperlink"

    auto_fit_columns(ws)

def main():
    ti_rows = []
    raw_rows = []

    vistos_ti = set()   # dedup TI
    vistos_raw = set()  # dedup RAW

    pagina = 1
    while True:
        params = {"dataFinal": DATA_FINAL, "pagina": pagina, "tamanhoPagina": TAMANHO_PAGINA}
        status, payload, real_url, msg = request_json(URL_PROPOSTA, params)

        if status == "END":
            print(f"ℹ Fim da paginação (204). Última URL: {real_url}")
            break

        if status != "OK":
            print(f"❌ Erro na API ({status}). URL: {real_url}")
            if msg:
                print("DETALHE:", msg)
            break

        itens = extract_list(payload)
        if not itens:
            print("ℹ Sem itens retornados. Encerrando.")
            break

        for it in itens:
            orgao = get_orgao(it)
            valor = get_valor(it)
            objeto = get_objeto(it)
            numero = get_numero_controle(it)

            # =========================
            # RAW (sem filtro)
            # =========================
            if len(raw_rows) < MAX_RAW:
                key_raw = numero.strip() if numero else (orgao + "|" + objeto + "|" + str(valor)).strip()
                if key_raw and key_raw not in vistos_raw:
                    vistos_raw.add(key_raw)
                    raw_rows.append({
                        "orgao": orgao,
                        "valor_estimado": valor,
                        "objeto": objeto,
                        "numeroControlePNCP": numero,
                        "link_pesquisa_pncp": pncp_search_link(numero),
                    })

            # =========================
            # TI (filtrado)
            # =========================
            if len(ti_rows) < MAX_TITENS_TI:
                if not is_ti(objeto):
                    continue

                key_ti = numero.strip() if numero else (orgao + "|" + objeto + "|" + str(valor)).strip()
                if key_ti and key_ti in vistos_ti:
                    continue
                if key_ti:
                    vistos_ti.add(key_ti)

                ti_rows.append({
                    "orgao": orgao,
                    "valor_estimado": valor,
                    "objeto": objeto,
                    "numeroControlePNCP": numero,
                    "link_pesquisa_pncp": pncp_search_link(numero),
                })

        # se já pegou tudo que precisa, para cedo
        if len(ti_rows) >= MAX_TITENS_TI and len(raw_rows) >= MAX_RAW:
            break

        pagina += 1

    # =========================
    # GERAR XLSX
    # =========================
    wb = Workbook()

    headers = ["orgao", "valor_estimado", "objeto", "numeroControlePNCP", "link_pesquisa_pncp"]

    ws1 = wb.active
    ws1.title = "proposta_aberta_TI"
    write_sheet(ws1, headers, ti_rows, link_col_name="link_pesquisa_pncp")

    ws2 = wb.create_sheet("proposta_aberta_RAW")
    write_sheet(ws2, headers, raw_rows, link_col_name="link_pesquisa_pncp")

    out = "licitacoes_TI_proposta_aberta.xlsx"
    wb.save(out)

    print(f"\n✅ Excel gerado: {out}")
    print(f"✅ Total TI encontrado: {len(ti_rows)}")
    print(f"✅ Total RAW salvo (amostra): {len(raw_rows)}")

    if len(ti_rows) == 0:
        print("⚠️ Observação: hoje não apareceu nenhuma 'proposta_aberta' que bateu no filtro de TI.")
        print("   Abra a aba 'proposta_aberta_RAW' para ver o que veio do PNCP.")
        print("   Se quiser, eu ajusto as palavras-chave para o seu tipo de serviço (ex.: software/SaaS).")

if __name__ == "__main__":
    main()