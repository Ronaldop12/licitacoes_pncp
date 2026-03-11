"""
========================================
RADAR DE LICITAÇÕES DE TI - PNCP
========================================
Sistema de monitoramento automático de licitações de TI
publicadas no Portal Nacional de Contratações Públicas

Requisitos:
- Python 3.10+
- requests, pandas, openpyxl

Autor: Sistema Automático PNCP
Data: 2026
========================================
"""

import requests
import pandas as pd
import os
import json
import time
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import logging
from historico_db import HistoricoDB
from fases_db import FasesDB

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

from constantes import PALAVRAS_TI, PALAVRAS_EXCLUSAO

# ==================== CONFIGURAÇÕES ====================

# URL da API PNCP
API_URL = "https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao"

# Arquivos de saída
OUTPUT_EXCEL = "radar_licitacoes_TI_PRO.xlsx"
OUTPUT_CSV = "dados/licitacoes.csv"
STATE_FILE = "radar_state.json"

CAMPOS_SAIDA = [
    "orgao",
    "cnpj_orgao",
    "objeto",
    "valor_estimado",
    "data_publicacao",
    "data_abertura",
    "data_encerramento",
    "uf",
    "municipio",
    "numero_edital",
    "modalidade",
    "status",
    "criterio_julgamento",
    "link_edital",
    "fonte",
]

# Parâmetros de performance
MAX_LICITACOES = 20000
TEMPO_ESPERA_ENTRE_REQUISICOES = 1  # segundos
TIMEOUT_REQUISICAO = 120  # segundos
MAX_TENTATIVAS = 5
DIAS_ATRAS = 15

# ==================== CLASSES ====================

class RadarLicitacoesTI:
    """
    Classe principal para coleta e processamento de licitações de TI
    """
    
    def __init__(self):
        self.dados_coletados: List[Dict] = []
        self.total_licitacoes = 0
        self.total_ti = 0
        self.total_duplicadas = 0
        self.falha_conectividade = False
        self.usando_cache = False
        self.erros = []
        self.sessao = self._criar_sessao_http()

    def _criar_sessao_http(self) -> requests.Session:
        """Cria sessao HTTP compartilhada."""
        sessao = requests.Session()
        sessao.headers.update(
            {
                "Accept": "application/json",
                "User-Agent": "RadarLicitacoesTI/1.1",
            }
        )
        return sessao

    @staticmethod
    def _erro_socket_permissao(mensagem_erro: str) -> bool:
        return "WinError 10013" in mensagem_erro
        
    def _fazer_requisicao(self, url: str, params: Dict) -> Optional[Dict]:
        """
        Realiza requisicao HTTP com retry automatico.

        Args:
            url: URL da API
            params: Parametros da requisicao

        Returns:
            Resposta JSON ou None em caso de erro.
        """
        for tentativa in range(MAX_TENTATIVAS):
            try:
                logger.info(f"Requisicao (tentativa {tentativa + 1}/{MAX_TENTATIVAS}): {url}")

                resposta = self.sessao.get(url, params=params, timeout=TIMEOUT_REQUISICAO)

                if resposta.status_code == 200:
                    logger.info("Requisicao bem-sucedida")
                    return resposta.json()

                if resposta.status_code == 204:
                    logger.info("Sem conteudo (fim da paginacao)")
                    return None

                if resposta.status_code == 500:
                    logger.warning("Erro 500 do servidor PNCP")
                    if tentativa < MAX_TENTATIVAS - 1:
                        espera = 5 * (tentativa + 1)
                        logger.info(f"Aguardando {espera}s antes de tentar novamente...")
                        time.sleep(espera)
                    continue

                if resposta.status_code == 400:
                    logger.error("Erro 400: Parametros invalidos")
                    return None

                logger.warning(f"Status HTTP {resposta.status_code}")

            except requests.Timeout:
                logger.warning(
                    f"Timeout na requisicao (tentativa {tentativa + 1}/{MAX_TENTATIVAS})"
                )
                if tentativa < MAX_TENTATIVAS - 1:
                    time.sleep(2)

            except requests.ConnectionError as e:
                mensagem = str(e)
                logger.warning(
                    f"Erro de conexao: {mensagem} (tentativa {tentativa + 1}/{MAX_TENTATIVAS})"
                )

                if self._erro_socket_permissao(mensagem):
                    self.falha_conectividade = True
                    self.erros.append(
                        "Bloqueio de socket detectado (WinError 10013). "
                        "Verifique firewall, antivirus, VPN/proxy e permissao de rede para python.exe."
                    )
                    logger.error(
                        "Bloqueio de rede detectado (WinError 10013). "
                        "Coleta online interrompida para evitar tentativas inuteis."
                    )
                    return None

                if tentativa < MAX_TENTATIVAS - 1:
                    time.sleep(2)

            except json.JSONDecodeError:
                logger.warning(
                    f"Erro ao decodificar JSON (tentativa {tentativa + 1}/{MAX_TENTATIVAS})"
                )

            except Exception as e:
                logger.error(f"Erro inesperado: {e}")
                self.erros.append(f"Tentativa {tentativa + 1}: {str(e)}")

        logger.error(f"Falha apos {MAX_TENTATIVAS} tentativas")
        return None
    def _eh_licitacao_ti(self, texto: str) -> bool:
        """
        Verifica se a licitação contém palavras-chave de TI
        com filtro de falsos positivos.
        """
        if not texto:
            return False
            
        texto_lower = str(texto).lower()
        
        # Verificar exclusões primeiro
        for exclusao in PALAVRAS_EXCLUSAO:
            if exclusao in texto_lower:
                return False
        
        for palavra in PALAVRAS_TI:
            if palavra in texto_lower:
                return True
        
        return False
    
    def _processar_licitacao(self, item: Dict) -> Optional[Dict]:
        """
        Processa e filtra uma licitação individual
        
        Args:
            item: Item da resposta da API
            
        Returns:
            Licitação processada ou None se não for TI
        """
        try:
            # Extrair campo principal: "objetoCompra" (não "objeto")
            objeto = item.get("objetoCompra", "")
            
            # Filtro de TI
            if not self._eh_licitacao_ti(objeto):
                return None
            
            # Extrair informações de órgão
            orgao_entidade = item.get("orgaoEntidade", {})
            unidade_orgao = item.get("unidadeOrgao", {})
            
            if isinstance(orgao_entidade, dict):
                nome_orgao = orgao_entidade.get("razaoSocial", "N/A")
            else:
                nome_orgao = str(orgao_entidade)
            
            if isinstance(unidade_orgao, dict):
                uf = unidade_orgao.get("ufSigla", "N/A")
                municipio = unidade_orgao.get("municipioNome", "N/A")
            else:
                uf = "N/A"
                municipio = "N/A"
            
            # Montar registro processado com campos expandidos
            numero_controle = item.get("numeroControlePNCP", "N/A")
            cnpj = ""
            if isinstance(orgao_entidade, dict):
                cnpj = orgao_entidade.get("cnpj", "")
            
            # Link direto para o edital no PNCP
            link = ""
            if numero_controle and numero_controle != "N/A":
                link = f"https://pncp.gov.br/app/editais/{numero_controle}"

            registro = {
                "orgao": nome_orgao,
                "cnpj_orgao": cnpj,
                "objeto": objeto,
                "valor_estimado": item.get("valorTotalEstimado", 0),
                "data_publicacao": item.get("dataPublicacaoPncp", "N/A"),
                "data_abertura": item.get("dataAberturaProposta", "N/A"),
                "data_encerramento": item.get("dataEncerramentoProposta", "N/A"),
                "uf": uf,
                "municipio": municipio,
                "numero_edital": numero_controle,
                "modalidade": item.get("modalidadeNome", "N/A"),
                "status": item.get("situacaoCompraNome", "N/A"),
                "criterio_julgamento": item.get("tipoCriterioJulgamentoNome", "N/A"),
                "link_edital": link,
                "fonte": "PNCP",
                "categoria_item": item.get("categoriaNome", "N/A"),
                "codigo_catmat_catser": item.get("codigoClasseItemMaterial", item.get("codigoGrupoMaterial", "N/A")),
            }
            
            return registro
            
        except Exception as e:
            logger.warning(f"Erro ao processar licitação: {e}")
            return None
    
    def coletar_licitacoes(self) -> List[Dict]:
        """
        Coleta licitacoes de TI com paginacao automatica.
        Utiliza estrategia de coleta por periodo e modalidades principais.

        Returns:
            Lista de licitacoes processadas.
        """
        logger.info("=" * 60)
        logger.info("INICIANDO COLETA DE LICITACOES DE TI")
        logger.info("=" * 60)

        data_fim = datetime.now()
        data_inicio = data_fim - timedelta(days=DIAS_ATRAS)

        logger.info(f"Periodo: {data_inicio.date()} a {data_fim.date()}")

        numeros_edital_coletados = set()
        # Todas as modalidades da API PNCP
        modalidades = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]

        for modalidade in modalidades:
            logger.info(f"\n--- COLETANDO MODALIDADE {modalidade} ---")
            pagina = 1
            paginas_vazias = 0

            while len(self.dados_coletados) < MAX_LICITACOES and paginas_vazias < 3:
                logger.info(f"Pagina {pagina} (Modalidade: {modalidade})")

                params = {
                    "dataInicial": data_inicio.strftime("%Y%m%d"),
                    "dataFinal": data_fim.strftime("%Y%m%d"),
                    "codigoModalidadeContratacao": modalidade,
                    "pagina": pagina,
                }

                resposta = self._fazer_requisicao(API_URL, params)

                if resposta is None:
                    if self.falha_conectividade:
                        logger.error("Coleta interrompida por falha de conectividade com a API.")
                    else:
                        logger.info("Resposta nula - fim da paginacao")
                    break

                dados_pagina = resposta.get("data", [])

                if not dados_pagina:
                    paginas_vazias += 1
                    logger.info(f"Pagina vazia ({paginas_vazias}/3)")
                    if paginas_vazias >= 3:
                        break
                    pagina += 1
                    time.sleep(TEMPO_ESPERA_ENTRE_REQUISICOES)
                    continue

                paginas_vazias = 0
                licitacoes_nesta_pagina = 0

                for item in dados_pagina:
                    if len(self.dados_coletados) >= MAX_LICITACOES:
                        logger.info(f"Limite de {MAX_LICITACOES} licitacoes atingido")
                        break

                    numero_edital = item.get("numeroControlePNCP", "")
                    if numero_edital in numeros_edital_coletados:
                        self.total_duplicadas += 1
                        continue

                    self.total_licitacoes += 1
                    licitacao_processada = self._processar_licitacao(item)

                    if licitacao_processada:
                        self.dados_coletados.append(licitacao_processada)
                        numeros_edital_coletados.add(numero_edital)
                        self.total_ti += 1
                        licitacoes_nesta_pagina += 1

                if len(self.dados_coletados) >= MAX_LICITACOES:
                    logger.info("Limite de licitacoes atingido")
                    break

                logger.info(f"  {licitacoes_nesta_pagina} licitacoes TI nesta pagina")
                logger.info(f"  Total: {self.total_ti} de {self.total_licitacoes}")

                time.sleep(TEMPO_ESPERA_ENTRE_REQUISICOES)
                pagina += 1

            if self.falha_conectividade:
                break

        logger.info("\n" + "=" * 60)
        logger.info("COLETA FINALIZADA")
        logger.info(f"Total de licitacoes verificadas: {self.total_licitacoes}")
        logger.info(f"Licitacoes de TI encontradas: {self.total_ti}")
        logger.info(f"Licitacoes duplicadas evitadas: {self.total_duplicadas}")
        logger.info("=" * 60)

        return self.dados_coletados

    def carregar_cache_csv(self, caminho: str = OUTPUT_CSV) -> bool:
        """Carrega dados de cache local quando a API estiver indisponivel."""
        if not os.path.exists(caminho):
            logger.error(f"Arquivo de cache nao encontrado: {caminho}")
            return False

        try:
            df = pd.read_csv(caminho, encoding="utf-8")
        except Exception as e:
            logger.error(f"Falha ao carregar cache CSV: {e}")
            self.erros.append(f"Cache CSV: {str(e)}")
            return False

        colunas_faltantes = [col for col in CAMPOS_SAIDA if col not in df.columns]
        if colunas_faltantes:
            logger.error(
                "Cache CSV invalido. Colunas ausentes: " + ", ".join(colunas_faltantes)
            )
            self.erros.append("Cache CSV invalido: colunas ausentes")
            return False

        df = df[CAMPOS_SAIDA].copy()
        df["valor_estimado"] = pd.to_numeric(df["valor_estimado"], errors="coerce").fillna(0)
        self.dados_coletados = df.to_dict("records")
        self.total_ti = len(self.dados_coletados)
        self.usando_cache = True
        logger.warning(
            f"API indisponivel. Usando cache local ({len(self.dados_coletados)} registros): {caminho}"
        )
        return True
    def _remover_duplicatas(self) -> None:
        """Remove registros duplicados baseado no número de edital (identificador único)"""
        logger.info("Removendo duplicatas...")
        
        if not self.dados_coletados:
            logger.warning("Nenhum dado para remover duplicatas")
            return
        
        df = pd.DataFrame(self.dados_coletados)
        tamanho_antes = len(df)
        
        # Usar numero_edital como chave única
        df = df.drop_duplicates(subset=["numero_edital"], keep="first")
        tamanho_depois = len(df)
        duplicatas_removidas = tamanho_antes - tamanho_depois
        
        if duplicatas_removidas > 0:
            logger.info(f"âœ“ {duplicatas_removidas} duplicatas removidas")
        
        self.dados_coletados = df.to_dict('records')
    
    def exportar_excel(self, caminho: str = OUTPUT_EXCEL) -> bool:
        """
        Exporta licitações para Excel com formatação profissional
        
        Args:
            caminho: Caminho do arquivo Excel
            
        Returns:
            True se bem-sucedido
        """
        try:
            logger.info(f"Exportando para Excel: {caminho}")
            
            if not self.dados_coletados:
                logger.warning("Nenhum dado para exportar")
                return False
            
            df = pd.DataFrame(self.dados_coletados)
            
            # Formatar valor
            df['valor_estimado'] = pd.to_numeric(df['valor_estimado'], errors='coerce').fillna(0)
            
            # Ordenação por valor
            df = df.sort_values('valor_estimado', ascending=False)
            
            # Exportar com formatação
            with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Licitações TI', index=False)
                
                # Adicionar formatação profissional
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                workbook = writer.book
                worksheet = writer.sheets['Licitações TI']
                
                # Estilo do header
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                # Aplicar ao header
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-ajustar largura e aplicar bordas
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # Formatar valores monetários
                        if cell.column_letter == get_column_letter(df.columns.get_loc('valor_estimado')+1) and cell.row > 1:
                            cell.number_format = '"R$ "#,##0.00'
                        
                        # Formatar datas
                        if cell.column_letter == get_column_letter(df.columns.get_loc('data_publicacao')+1) if 'data_publicacao' in df.columns else False:
                            if cell.row > 1:
                                cell.number_format = 'dd/mm/yyyy'
                
                # Auto-ajustar largura
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Congelar header
                worksheet.freeze_panes = "A2"
            
            logger.info(f"âœ“ Excel exportado com sucesso: {len(df)} registros")
            return True
            
        except Exception as e:
            logger.error(f"âœ— Erro ao exportar Excel: {e}")
            self.erros.append(f"Exportação Excel: {str(e)}")
            return False
    
    def exportar_csv(self, caminho: str = OUTPUT_CSV) -> bool:
        """
        Exporta licitações para CSV
        
        Args:
            caminho: Caminho do arquivo CSV
            
        Returns:
            True se bem-sucedido
        """
        try:
            logger.info(f"Exportando para CSV: {caminho}")
            
            # Criar diretório se não existir
            diretorio = os.path.dirname(caminho)
            if diretorio and not os.path.exists(diretorio):
                os.makedirs(diretorio)
                logger.info(f"Diretório criado: {diretorio}")
            
            if not self.dados_coletados:
                logger.warning("Nenhum dado para exportar")
                return False
            
            df = pd.DataFrame(self.dados_coletados)
            df.to_csv(caminho, index=False, encoding='utf-8')
            
            logger.info(f"âœ“ CSV exportado com sucesso: {len(df)} registros")
            return True
            
        except Exception as e:
            logger.error(f"âœ— Erro ao exportar CSV: {e}")
            self.erros.append(f"Exportação CSV: {str(e)}")
            return False
    
    def gerar_relatorio(self) -> Dict:
        """
        Gera relatório estatístico
        
        Returns:
            Dicionário com estatísticas
        """
        if not self.dados_coletados:
            return {
                "total": 0,
                "orgaos": 0,
                "maior_valor": 0,
                "valor_total": 0,
                "valor_medio": 0,
                "estados": 0,
                "top_orgaos": {},
                "distribuicao_ufs": {}
            }
        
        df = pd.DataFrame(self.dados_coletados)
        
        relatorio = {
            "total": len(df),
            "orgaos": df['orgao'].nunique(),
            "maior_valor": df['valor_estimado'].max(),
            "valor_total": df['valor_estimado'].sum(),
            "valor_medio": df['valor_estimado'].mean(),
            "estados": df['uf'].nunique(),
            "top_orgaos": df['orgao'].value_counts().head(10).to_dict(),
            "distribuicao_ufs": df['uf'].value_counts().to_dict(),
        }
        
        return relatorio
    
    def salvar_estado(self, caminho: str = STATE_FILE) -> None:
        """Salva estado da coleta em JSON"""
        try:
            estado = {
                "data_execucao": datetime.now().isoformat(),
                "total_licitacoes": self.total_licitacoes,
                "total_ti": self.total_ti,
                "total_duplicadas": self.total_duplicadas,
                "total_coletado": len(self.dados_coletados),
                "falha_conectividade": self.falha_conectividade,
                "usando_cache": self.usando_cache,
                "erros": self.erros
            }
            
            with open(caminho, 'w', encoding='utf-8') as f:
                json.dump(estado, f, indent=4, ensure_ascii=False)
            
            logger.info(f"Estado salvo em {caminho}")
        except Exception as e:
            logger.error(f"Erro ao salvar estado: {e}")


def main():
    """Funcao principal"""
    try:
        radar = RadarLicitacoesTI()

        radar.coletar_licitacoes()

        # Se houver bloqueio de rede, tenta operar com cache local.
        if not radar.dados_coletados and radar.falha_conectividade:
            logger.warning(
                "Nenhum dado novo coletado por falha de conectividade. Tentando cache local..."
            )
            radar.carregar_cache_csv()

        radar._remover_duplicatas()

        # Exporta apenas quando houve coleta online de dados novos.
        if radar.dados_coletados and not radar.usando_cache:
            radar.exportar_excel()
            radar.exportar_csv()

            # Registrar no histórico versionado
            try:
                historico = HistoricoDB()
                df_hist = pd.DataFrame(radar.dados_coletados)
                historico.registrar_coleta(
                    df_hist, fonte="PNCP",
                    total_verificadas=radar.total_licitacoes,
                    total_duplicadas=radar.total_duplicadas,
                )
                logger.info("Coleta registrada no histórico versionado")
            except Exception as e:
                logger.warning(f"Falha ao registrar histórico: {e}")

            # Rastrear mudanças de status/fases
            try:
                fases = FasesDB()
                df_fases = pd.DataFrame(radar.dados_coletados)
                mudancas = fases.processar_coleta(df_fases)
                if mudancas:
                    logger.info(f"Detectadas {len(mudancas)} mudanças de status")
                    for m in mudancas[:5]:
                        logger.info(f"  {m['numero_edital']}: {m['status_anterior']} → {m['status_novo']}")
            except Exception as e:
                logger.warning(f"Falha ao rastrear fases: {e}")

            # Registrar preços por categoria CATMAT/CATSER
            try:
                from precos_db import PrecosDB
                precos = PrecosDB()
                df_precos = pd.DataFrame(radar.dados_coletados)
                precos.registrar_precos(df_precos)
                logger.info("Preços registrados por categoria")
            except Exception as e:
                logger.warning(f"Falha ao registrar preços: {e}")

            # Registrar métricas da coleta
            try:
                from metricas import MetricasDB
                met = MetricasDB()
                met.registrar_coleta(
                    total_ti=len(radar.dados_coletados),
                    total_verificadas=radar.total_licitacoes,
                    duracao_s=0,
                    erros=len(radar.erros),
                )
                logger.info("Métricas da coleta registradas")
            except Exception as e:
                logger.warning(f"Falha ao registrar métricas: {e}")

        relatorio = radar.gerar_relatorio()

        logger.info("\n" + "=" * 60)
        logger.info("RESUMO DA EXECUCAO")
        logger.info("=" * 60)
        logger.info(f"Total de licitacoes coletadas: {relatorio['total']}")
        logger.info(f"Total de orgaos: {relatorio['orgaos']}")
        logger.info(f"Total de estados: {relatorio['estados']}")
        logger.info(f"Valor total (R$): {relatorio['valor_total']:,.2f}")
        logger.info(f"Maior licitacao (R$): {relatorio['maior_valor']:,.2f}")

        if relatorio['top_orgaos']:
            logger.info("\nTop 5 orgaos que mais contratam TI:")
            for i, (orgao, qtd) in enumerate(list(relatorio['top_orgaos'].items())[:5], 1):
                logger.info(f"  {i}. {orgao}: {qtd} licitacoes")

        logger.info("=" * 60)

        radar.salvar_estado()

        if radar.usando_cache:
            logger.warning(
                "Execucao concluida em modo cache: API indisponivel, dashboard atualizado com dados locais."
            )
        elif radar.falha_conectividade and not radar.dados_coletados:
            raise RuntimeError(
                "Falha de conectividade com a API PNCP e cache local indisponivel. Verifique rede/firewall."
            )
        else:
            logger.info("\nEXECUCAO CONCLUIDA COM SUCESSO\n")

    except Exception as e:
        logger.error(f"Erro fatal: {e}")
        raise

if __name__ == "__main__":
    main()
